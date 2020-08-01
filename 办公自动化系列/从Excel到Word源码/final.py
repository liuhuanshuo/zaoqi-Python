'''
公众号：早起Python
制作：陈熹
兴趣范围：生物信息 / 数据分析 / 网络爬虫
简书：半为花间酒
Email：chenx6542@foxmail.com
'''


from openpyxl import load_workbook
from docx import Document
import datetime
import os


def GetDesktopPath():
    return os.path.join(os.path.expanduser("~"), 'Desktop')


path = GetDesktopPath() + '/资料/'

workbook = load_workbook(filename=path + '数据.xlsx')
sheet = workbook.active

# 获取有数据范围
# print(sheet.dimensions)
# A1:W10

# SQE
SQE = sheet['Q2'].value

# 供应商&制造商
supplier = sheet['G2'].value

# 采购单号
C2_10 = sheet['C2:C10']  # 返回cell.tuple对象
vC2_10 = [str(cell[0].value) for cell in C2_10]
order_num = ','.join(set(vC2_10))
order_num_title = '&'.join(set(vC2_10))

# 产品型号
T2_10 = sheet['T2:T10']
vT2_10 = [str(cell[0].value) for cell in T2_10]
ptype = ','.join(set(vT2_10))

# 产品描述
P2_10 = sheet['P2:P10']
vP2_10 = [str(cell[0].value) for cell in P2_10]
info = ','.join(set(vP2_10))
info_title = '&'.join(set(vP2_10))

# 日期
today = datetime.datetime.today()
time = today.strftime('%Y年%m月%d日')

# 验货数量
V2_10 = sheet['V2:V10']
vV2_10 = [int(cell[0].value) for cell in V2_10]
total_num = sum(vV2_10)  # 计算总数量

# 验货箱数
W2_10 = sheet['W2:W10']
vW2_10 = [int(cell[0].value) for cell in W2_10]
box_num = sum(vW2_10)

title = f'{order_num_title}-{supplier}-{total_num}-{info_title}-{time}-验货报告'

print(title)

doc_path = path + '模板.doc'
docx_path = doc_path + 'x'

# doc转docx的函数

document = Document(docx_path)

# 读取word中的所有表格
tables = document.tables
# print(len(tables))
# 15

# 开始填表
tables[0].cell(1, 1).text = SQE

tables[1].cell(1, 1).text = supplier
tables[1].cell(2, 1).text = supplier
tables[1].cell(3, 1).text = ptype
tables[1].cell(4, 1).text = info
tables[1].cell(5, 1).text = order_num
tables[1].cell(7, 1).text = time

for i in range(2, 11):
    tables[6].cell(i, 0).text = str(sheet[f'T{i}'].value)
    tables[6].cell(i, 1).text = str(sheet[f'P{i}'].value)
    tables[6].cell(i, 2).text = str(sheet[f'C{i}'].value)
    tables[6].cell(i, 4).text = str(sheet[f'V{i}'].value)
    tables[6].cell(i, 5).text = str(sheet[f'V{i}'].value)
    tables[6].cell(i, 6).text = '0'
    tables[6].cell(i, 7).text = str(sheet[f'W{i}'].value)
    tables[6].cell(i, 8).text = '0'

tables[6].cell(12, 4).text = str(total_num)
tables[6].cell(12, 5).text = str(total_num)
tables[6].cell(12, 7).text = str(box_num)

for i in range(2, 11):
    tables[13].cell(i - 1, 0).text = str(sheet[f'T{i}'].value)
    tables[13].cell(i - 1, 1).text = str(sheet[f'U{i}'].value)
    tables[13].cell(i - 1, 2).text = str(sheet[f'U{i}'].value)
    tables[13].cell(i - 1, 3).text = str(sheet[f'U{i}'].value)

os.remove(docx_path)

document.save(path + f'{title}.docx')
print('文件已生成，文件运行')
