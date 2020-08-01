'''
公众号：早起Python
作者：陈熹
请在桌面创建一个文件夹并命名为data，然后将测试Excel文件放入data文件夹中！
'''
from openpyxl import load_workbook
import os
import glob
import random
import pandas as pd
import re
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border
from openpyxl.styles import Font

def GetDesktopPath():
    return os.path.join(os.path.expanduser("~"), 'Desktop')

path = glob.glob(f'{GetDesktopPath()}/data/*.xls*')[0]
workbook = load_workbook(filename=path)
sheet_init = workbook.active
name_lst = ['皮卡丘', '小火龙', '杰尼龟', '妙蛙种子', '风速狗', '小拳石', '飞天螳螂']
place_lst = [chr(i).upper() for i in range(97, 123)]
activity_lst = ['椭圆机', '篮球', '足球', '羽毛球', '跳绳']
source_lst = ['朋友介绍', '微信聊天', '网页弹窗', '其他']
for i in range(30):
    sheet = workbook.copy_worksheet(sheet_init)
    sheet.title = f'{i+1}日'
    for j in range(random.randint(10, 30)):
        for row in sheet.iter_rows(min_row=3+j, max_row=3+j):
            info = [f'{j+1}', f'{i+1}日', f'{random.choice(name_lst)}', f'{random.choice(place_lst)}馆',
                    f'{random.choice(activity_lst)}', f'{random.choice(source_lst)}', f'{random.randint(1, 10)}',
                    '无', f'{random.choice(["Y", "N"])}', f'{random.choice(["Y", "N"])}', f'{random.choice(["Y", "N"])}']
            for index, k in enumerate(info):
                row[index].value = k

workbook.save(filename=f'{GetDesktopPath()}/data/results.xlsx')

path_new = glob.glob(f'{GetDesktopPath()}/data/results.xls*')[0]
# 方便获取总表数便于遍历
workbook = load_workbook(path_new)
sheetnames = workbook.sheetnames
df_lst = []
for i in range(1, len(sheetnames)):
    df = pd.read_excel(path_new, encoding='utf-8', sheet_name=i, skiprows=1)
    df_lst.append(df)

df_total = pd.concat(df_lst, axis=0, ignore_index=True)
df_total['编号'] = df_total.index + 1

writer = pd.ExcelWriter(path_new, engine='openpyxl')
writer.book = workbook
workbook.remove(workbook['汇总表'])
df_total.to_excel(excel_writer=writer, sheet_name=u'汇总表', index=None)
writer.close()
workbook._sheets.insert(0, workbook._sheets.pop())

sheet = workbook[sheetnames[0]]
sheet.insert_rows(idx=0)
font = Font(name='宋体', size=18, bold=True)
sheet['A1'] = '皮卡丘体育2020年06月新学员信息登记表'
sheet['A1'].font = font

req = ':(\w)'
weight = re.findall(req, sheet.dimensions)[0]
sheet.merge_cells(f'A1:{weight}1')

alignment = Alignment(horizontal='center', vertical='center')
side = Side(style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)

rows = sheet[f'{sheet.dimensions}']
for row in rows:
    for cell in row:
        cell.alignment = alignment
        cell.border = border

sheet.row_dimensions[1].height = 38
sheet.row_dimensions[2].height = 38

letter_lst = [chr(i+64).upper() for i in range(2, ord(weight)-ord('A')+1+1)]
sheet.column_dimensions['A'].width = 8
for i in letter_lst:
     sheet.column_dimensions[f'{i}'].width = 14

workbook.save(filename=f'{GetDesktopPath()}/data/results.xlsx')
print('文件已生成')