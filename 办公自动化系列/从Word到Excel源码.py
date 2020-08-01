from docx import Document
import datetime
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
header = ['序号', '收文时间', '办文编号', '文件标题', '文号', '备注']
sheet.append(header)


path = r'C:\Users\word.docx'
document = Document(path)
tables = document.tables

n = 0
for j in range(len(tables)):
    for i in range(0, len(tables[j].rows) + 1, 3):
        try:
            # 日期
            date = tables[j].cell(i, 1).text
            if '/' in date:
                date = datetime.datetime.strptime(
                    date, '%d/%m').strftime('2020-%m-%d')
            else:
                date = '-'
            # 标题
            title = tables[j].cell(i + 1, 1).text.strip()
            # 文号
            dfn = tables[j].cell(i, 3).text.strip()
            n += 1
            print(n, date, title, dfn)
            row = [n, date, ' ', title, dfn, ' ']
            sheet.append(row)
        except Exception as error:
            # 捕获异常，也可以用log写到日志里方便查看和管理
            print(error)
            continue

wb.save(r'C:\Users\20200420.xlsx')
