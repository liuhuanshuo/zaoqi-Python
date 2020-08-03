from openpyxl import load_workbook, Workbook

# 从桌面上获取总表
filepath = r'C:\Users\chenx\Desktop\台账.xlsm'  # 根据实际情况进行修改
workbook = load_workbook(filepath)
# 创建新的Excel工作簿获取到工作表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 给新表写入表头
new_headers = ['名称', '配置', '提交日期', '受限操作', '操作时间', '状态', '存储位置']
new_sheet.append(new_headers)


for i in workbook.sheetnames:
    sheet = workbook[i]
    names = sheet['A']
    flag = 0
    for cell in names:
        if cell.value == keyword:
            flag = cell.row
            break
    if flag:   # 如果flag没有被修改则不需要顺序进行下列代码
        data_lst = []
        for cell in sheet[flag]:
            # 这里加上一个对内容的判断，是让无内容的行直接放空，而不是写入一个 none
            if cell.value:
                data_lst.append(str(cell.value))
            else:
                data_lst.append(' ')
        new_sheet.append(data_lst)


new_workbook.save(r'C:\Users\chenx\Desktop\台账查询.xlsx')
